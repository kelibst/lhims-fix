#!/usr/bin/env python3
"""
LHIMS Master Patient List Generator

Scans all extracted Excel files and generates a deduplicated list
of all unique patient numbers for JSON extraction.

Usage:
    python scripts/generate-patient-list.py

Output:
    master-patient-list.txt - Deduplicated list of patient numbers
"""

import openpyxl
import os
from pathlib import Path
import re

# Data folders to scan
DATA_FOLDERS = [
    'data/opd-register',
    'data/ipd-morbidity-mortality',
    'data/anc-register',
    'data/consulting-room',
    'data/medical-laboratory',
]

# Patient number pattern (XX-A01-AAANNNN) - Accepts all facility codes
PATIENT_NO_PATTERN = re.compile(r'[A-Z]{2}-[A-Z]\d{2}-[A-Z]{3}\d+')

def find_header_row(sheet, max_search=20):
    """Find the row with column headers"""
    for row_idx in range(1, min(max_search + 1, sheet.max_row + 1)):
        row = sheet[row_idx]
        non_empty = sum(1 for cell in row if cell.value)
        if non_empty >= 5:  # Row with at least 5 non-empty cells
            return row_idx
    return None

def find_patient_no_column(headers):
    """Find the Patient No. column index"""
    for i, header in enumerate(headers):
        if header and 'patient' in header.lower() and 'no' in header.lower():
            return i
    return None

def extract_patients_from_excel(filepath):
    """Extract all patient numbers from an Excel file"""
    patients = set()

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active

        # Find header row
        header_row_idx = find_header_row(sheet)
        if not header_row_idx:
            print(f"  ⚠ No header row found in {os.path.basename(filepath)}")
            wb.close()
            return patients

        # Get headers
        headers = []
        for cell in sheet[header_row_idx]:
            val = str(cell.value).strip() if cell.value else ''
            headers.append(val)

        # Find Patient No. column
        patient_col_idx = find_patient_no_column(headers)
        if patient_col_idx is None:
            print(f"  ⚠ No 'Patient No.' column found in {os.path.basename(filepath)}")
            wb.close()
            return patients

        # Extract patient numbers
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row = sheet[row_idx]
            if patient_col_idx < len(row):
                cell_value = row[patient_col_idx].value
                if cell_value:
                    patient_no = str(cell_value).strip()
                    # Validate format
                    if PATIENT_NO_PATTERN.match(patient_no):
                        patients.add(patient_no)

        wb.close()

    except Exception as e:
        print(f"  ✗ Error reading {os.path.basename(filepath)}: {str(e)}")

    return patients

def generate_patient_list():
    """Generate master patient list from all Excel files"""
    print('='*70)
    print('LHIMS MASTER PATIENT LIST GENERATOR')
    print('='*70)
    print('')

    all_patients = set()
    file_count = 0

    # Process each data folder
    for folder in DATA_FOLDERS:
        folder_path = Path(folder)

        if not folder_path.exists():
            print(f'⚠ Folder not found: {folder}')
            continue

        print(f'📂 Scanning: {folder}')

        # Get all Excel files
        excel_files = list(folder_path.glob('*.xlsx')) + list(folder_path.glob('*.xls'))
        excel_files = [f for f in excel_files if not f.name.startswith('~$')]  # Exclude temp files

        folder_patients = set()

        for i, excel_file in enumerate(excel_files, 1):
            print(f'  [{i}/{len(excel_files)}] Processing: {excel_file.name}...', end='', flush=True)
            patients = extract_patients_from_excel(str(excel_file))
            folder_patients.update(patients)
            file_count += 1
            print(f' ✓ ({len(patients)} patients)')

        print(f'  ✓ Found {len(folder_patients)} unique patients in {len(excel_files)} files')
        all_patients.update(folder_patients)
        print('')

    # Sort patients
    sorted_patients = sorted(list(all_patients))

    # Generate output file
    print('='*70)
    print('GENERATING OUTPUT FILE')
    print('='*70)
    print('')

    output_lines = [
        '# LHIMS Master Patient List',
        '# Generated automatically from extracted Excel files',
        '#',
        f'# Generated on: {import_datetime()}',
        f'# Total unique patients: {len(sorted_patients)}',
        f'# Files scanned: {file_count}',
        f'# Source folders: {", ".join(DATA_FOLDERS)}',
        '#',
        '# Format: XX-A01-AAANNNN (one patient number per line, all facility codes)',
        '# Lines starting with # are comments and will be ignored',
        '#',
        ''
    ] + sorted_patients

    output_file = 'master-patient-list.txt'
    with open(output_file, 'w') as f:
        f.write('\n'.join(output_lines))

    print(f'✓ Master patient list generated: {output_file}')
    print(f'  Total unique patients: {len(sorted_patients):,}')
    print(f'  Files scanned: {file_count}')
    print('')

    # Show sample
    print('Sample patients:')
    for patient in sorted_patients[:10]:
        print(f'  - {patient}')
    if len(sorted_patients) > 10:
        print(f'  ... and {len(sorted_patients) - 10} more')
    print('')

    print('='*70)
    print('NEXT STEPS')
    print('='*70)
    print('1. Review: master-patient-list.txt')
    print('2. Run extraction: npm run extract:patients')
    print('3. Find JSON data in: data/patient-json/')
    print('')

def import_datetime():
    """Import datetime lazily"""
    from datetime import datetime
    return datetime.now().isoformat()

if __name__ == '__main__':
    generate_patient_list()
